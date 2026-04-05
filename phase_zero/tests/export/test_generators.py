"""
Generator Tests — CSV (CSV-01→05), Excel (XLS-01→05), Power BI (PBI-01→05).

All three generators must import from EXPORT_COLUMN_ORDER.
Tests verify column order, content, and format correctness.
"""

import csv
import json
import pytest
from pathlib import Path

from app.export.csv_generator import generate_csv
from app.export.excel_generator import generate_excel
from app.export.power_bi_generator import generate_power_bi
from app.export.column_order import EXPORT_COLUMN_ORDER


# Standard test row
MOCK_ROWS = [
    {
        "region": "us-east-1",
        "gpu_pool_id": "pool-a",
        "date": "2026-01-15",
        "billing_period": "2026-01",
        "allocation_target": "tenant-A",
        "unallocated_type": "",
        "failed_tenant_id": "",
        "gpu_hours": "100.000000",
        "cost_per_gpu_hour": "2.500000",
        "contracted_rate": "4.000000",
        "revenue": "400.00",
        "cogs": "250.00",
        "gross_margin": "150.00",
        "session_id": "test-sess-id",
        "source_files": '["telemetry.csv","cost.csv"]',
    },
    {
        "region": "us-east-1",
        "gpu_pool_id": "pool-a",
        "date": "2026-01-15",
        "billing_period": "2026-01",
        "allocation_target": "unallocated",
        "unallocated_type": "identity_broken",
        "failed_tenant_id": "tenant-BROKEN",
        "gpu_hours": "50.000000",
        "cost_per_gpu_hour": "2.500000",
        "contracted_rate": None,
        "revenue": "0.00",
        "cogs": "125.00",
        "gross_margin": "-125.00",
        "session_id": "test-sess-id",
        "source_files": '["telemetry.csv","cost.csv"]',
    },
]


# ── CSV Generator ─────────────────────────────────────────────────


class TestCSVGenerator:
    """Step 7.6 — CSV Generator."""

    # CSV-01: file created with .csv extension
    def test_csv_01_file_created(self, tmp_path):
        path = generate_csv(MOCK_ROWS, tmp_path, "sess-1")
        assert path.exists()
        assert path.suffix == ".csv"

    # CSV-02: header matches EXPORT_COLUMN_ORDER
    def test_csv_02_header_matches_column_order(self, tmp_path):
        path = generate_csv(MOCK_ROWS, tmp_path, "sess-1")
        with open(path, "r") as f:
            reader = csv.reader(f)
            headers = next(reader)
        assert headers == list(EXPORT_COLUMN_ORDER)

    # CSV-03: correct row count (data rows only)
    def test_csv_03_row_count(self, tmp_path):
        path = generate_csv(MOCK_ROWS, tmp_path, "sess-1")
        with open(path, "r") as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            data_rows = list(reader)
        assert len(data_rows) == 2

    # CSV-04: None values written as empty string
    def test_csv_04_none_as_empty(self, tmp_path):
        path = generate_csv(MOCK_ROWS, tmp_path, "sess-1")
        with open(path, "r") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        # Row 2 has contracted_rate = None
        assert rows[1]["contracted_rate"] == ""

    # CSV-05: session_id and source_files in output
    def test_csv_05_metadata_in_output(self, tmp_path):
        path = generate_csv(MOCK_ROWS, tmp_path, "sess-1")
        with open(path, "r") as f:
            reader = csv.DictReader(f)
            row = next(reader)
        assert row["session_id"] == "test-sess-id"
        assert "telemetry.csv" in row["source_files"]


# ── Excel Generator ───────────────────────────────────────────────


class TestExcelGenerator:
    """Step 7.7 — Excel Generator."""

    # XLS-01: file created with .xlsx extension
    def test_xls_01_file_created(self, tmp_path):
        path = generate_excel(MOCK_ROWS, tmp_path, "sess-1")
        assert path.exists()
        assert path.suffix == ".xlsx"

    # XLS-02: header matches EXPORT_COLUMN_ORDER
    def test_xls_02_header_matches_column_order(self, tmp_path):
        from openpyxl import load_workbook

        path = generate_excel(MOCK_ROWS, tmp_path, "sess-1")
        wb = load_workbook(str(path), read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        wb.close()
        assert headers == list(EXPORT_COLUMN_ORDER)

    # XLS-03: correct row count
    def test_xls_03_row_count(self, tmp_path):
        from openpyxl import load_workbook

        path = generate_excel(MOCK_ROWS, tmp_path, "sess-1")
        wb = load_workbook(str(path), read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        # First row is header, rest are data
        assert len(all_rows) - 1 == 2

    # XLS-04: None values written as empty string
    def test_xls_04_none_as_empty(self, tmp_path):
        from openpyxl import load_workbook

        path = generate_excel(MOCK_ROWS, tmp_path, "sess-1")
        wb = load_workbook(str(path), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        headers = list(rows[0])
        cr_idx = headers.index("contracted_rate")
        # Row 2 (index 2 in all_rows) has None contracted_rate
        # openpyxl normalizes "" back to None on read_only read-back
        assert rows[2][cr_idx] in (None, "")

    # XLS-05: sheet name is "GPU Margin Export"
    def test_xls_05_sheet_name(self, tmp_path):
        from openpyxl import load_workbook

        path = generate_excel(MOCK_ROWS, tmp_path, "sess-1")
        wb = load_workbook(str(path), read_only=True)
        assert wb.active.title == "GPU Margin Export"
        wb.close()


# ── Power BI Generator ────────────────────────────────────────────


class TestPowerBIGenerator:
    """Step 7.8 — Power BI Generator."""

    # PBI-01: file created with .txt extension
    def test_pbi_01_file_created(self, tmp_path):
        path = generate_power_bi(MOCK_ROWS, tmp_path, "sess-1")
        assert path.exists()
        assert path.suffix == ".txt"

    # PBI-02: pipe-delimited header matches EXPORT_COLUMN_ORDER
    def test_pbi_02_header_matches_column_order(self, tmp_path):
        path = generate_power_bi(MOCK_ROWS, tmp_path, "sess-1")
        with open(path, "r") as f:
            reader = csv.reader(f, delimiter="|")
            headers = next(reader)
        assert headers == list(EXPORT_COLUMN_ORDER)

    # PBI-03: correct row count
    def test_pbi_03_row_count(self, tmp_path):
        path = generate_power_bi(MOCK_ROWS, tmp_path, "sess-1")
        with open(path, "r") as f:
            reader = csv.reader(f, delimiter="|")
            next(reader)  # skip header
            data_rows = list(reader)
        assert len(data_rows) == 2

    # PBI-04: pipe delimiter used (not comma)
    def test_pbi_04_pipe_delimiter(self, tmp_path):
        path = generate_power_bi(MOCK_ROWS, tmp_path, "sess-1")
        with open(path, "r") as f:
            header_line = f.readline()
        assert "|" in header_line
        # Should not have unquoted commas as delimiters
        # (source_files JSON has commas but inside quotes)

    # PBI-05: source_files present in pipe-delimited output
    def test_pbi_05_source_files_present(self, tmp_path):
        path = generate_power_bi(MOCK_ROWS, tmp_path, "sess-1")
        with open(path, "r") as f:
            reader = csv.DictReader(f, delimiter="|")
            row = next(reader)
        assert "telemetry.csv" in row["source_files"]
