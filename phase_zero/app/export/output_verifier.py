"""
Output Verifier — Component 8/9.

Layer: Export.

6 checks on the generated export file:
  1. File exists
  2. Row count matches final.allocation_result
  3. Grain columns present (imported EXPORT_COLUMN_ORDER)
  4. Subtypes correct (Type A / identity_broken / capacity_idle)
  5. File is readable (no corruption)
  6. Metadata format (session_id + source_files as last two columns)

Spec: build-checklist.md — Step 7.9
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Literal

from pydantic import BaseModel

from app.export.column_order import EXPORT_COLUMN_ORDER, METADATA_COLUMNS


class VerificationResult(BaseModel):
    """Result of the 6-check output verification."""

    check_1_file_exists: bool = False
    check_2_row_count_match: bool = False
    check_3_columns_present: bool = False
    check_4_subtypes_correct: bool = False
    check_5_file_readable: bool = False
    check_6_metadata_format: bool = False
    errors: list[str] = []

    @property
    def all_passed(self) -> bool:
        return (
            self.check_1_file_exists
            and self.check_2_row_count_match
            and self.check_3_columns_present
            and self.check_4_subtypes_correct
            and self.check_5_file_readable
            and self.check_6_metadata_format
        )


VALID_SUBTYPES = {None, "", "identity_broken", "capacity_idle"}


def verify_output(
    filepath: Path,
    expected_row_count: int,
    file_format: Literal["csv", "excel", "power_bi"],
) -> VerificationResult:
    """
    Run 6 verification checks on a generated export file.

    Parameters
    ----------
    filepath : Path
        The generated file to verify.
    expected_row_count : int
        Number of data rows expected (from final.allocation_result).
    file_format : str
        "csv", "excel", or "power_bi".

    Returns
    -------
    VerificationResult
        All 6 check results + error messages.
    """
    result = VerificationResult()

    # ── Check 1: File exists ──────────────────────────────────────
    if not filepath.exists():
        result.errors.append(f"Check 1 FAIL: File not found: {filepath}")
        return result
    result.check_1_file_exists = True

    # ── Check 5: File is readable (attempt parse) ─────────────────
    try:
        headers, data_rows = _read_file(filepath, file_format)
    except Exception as exc:
        result.errors.append(f"Check 5 FAIL: File unreadable: {exc}")
        return result
    result.check_5_file_readable = True

    # ── Check 3: Grain columns present ────────────────────────────
    expected_cols = set(EXPORT_COLUMN_ORDER)
    actual_cols = set(headers)
    missing = expected_cols - actual_cols
    if missing:
        result.errors.append(
            f"Check 3 FAIL: Missing columns: {sorted(missing)}"
        )
    else:
        result.check_3_columns_present = True

    # ── Check 6: Metadata format (last two columns) ───────────────
    if len(headers) >= 2:
        last_two = headers[-2:]
        if last_two == list(METADATA_COLUMNS):
            result.check_6_metadata_format = True
        else:
            result.errors.append(
                f"Check 6 FAIL: Last two columns are {last_two}, "
                f"expected {list(METADATA_COLUMNS)}"
            )
    else:
        result.errors.append("Check 6 FAIL: Fewer than 2 columns")

    # ── Check 2: Row count matches ────────────────────────────────
    actual_count = len(data_rows)
    if actual_count == expected_row_count:
        result.check_2_row_count_match = True
    else:
        result.errors.append(
            f"Check 2 FAIL: Expected {expected_row_count} rows, "
            f"got {actual_count}"
        )

    # ── Check 4: Subtypes correct ─────────────────────────────────
    if result.check_3_columns_present and "unallocated_type" in headers:
        col_idx = headers.index("unallocated_type")
        invalid_subtypes = set()
        for row in data_rows:
            val = row[col_idx] if col_idx < len(row) else ""
            if val not in VALID_SUBTYPES:
                invalid_subtypes.add(val)
        if invalid_subtypes:
            result.errors.append(
                f"Check 4 FAIL: Invalid subtypes: {sorted(invalid_subtypes)}"
            )
        else:
            result.check_4_subtypes_correct = True
    else:
        # If columns aren't present, subtypes can't be checked
        result.check_4_subtypes_correct = result.check_3_columns_present

    return result


def _read_file(
    filepath: Path,
    file_format: Literal["csv", "excel", "power_bi"],
) -> tuple[list[str], list[list[str]]]:
    """
    Read headers and data rows from the export file.

    Returns (headers, data_rows) where data_rows is a list of lists.
    """
    if file_format == "excel":
        return _read_excel(filepath)

    # CSV and Power BI are both delimited text
    delimiter = "|" if file_format == "power_bi" else ","

    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=delimiter)
        headers = next(reader)
        data_rows = [row for row in reader]

    return headers, data_rows


def _read_excel(filepath: Path) -> tuple[list[str], list[list[str]]]:
    """Read headers and data from an Excel file."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl required for Excel verification")

    wb = load_workbook(str(filepath), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    headers = [str(c) if c is not None else "" for c in next(rows_iter)]
    data_rows = [
        [str(c) if c is not None else "" for c in row]
        for row in rows_iter
    ]
    wb.close()
    return headers, data_rows
